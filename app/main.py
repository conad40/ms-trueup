"""MS License True-Up — Unified API + Collector + Update System.

Single FastAPI service that:
- Serves the React frontend as static files
- Provides all API endpoints (hosts, compliance, settings, etc.)
- Runs scanner modules as background tasks
- Supports git-based self-updates from the web UI
"""
import os, math, io, re, glob, subprocess, asyncio, logging, secrets, hashlib
from contextlib import asynccontextmanager
from datetime import date, datetime, timezone

def _parse_date(v):
    """Convert a date string to a date object, or None."""
    if not v: return None
    if isinstance(v, date): return v
    return date.fromisoformat(v)
from pathlib import Path

import asyncpg
from cryptography.fernet import Fernet
from fastapi import FastAPI, Query, Header, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse, PlainTextResponse
from fastapi.staticfiles import StaticFiles
from typing import Optional

log = logging.getLogger("trueup")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(message)s")

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_PORT = int(os.environ.get("DB_PORT", "5432"))
DB_NAME = os.environ.get("DB_NAME", "trueup")
DB_USER = os.environ.get("DB_USER", "trueup")
DB_PASS = os.environ.get("DB_PASS", "trueup")
pool: asyncpg.Pool = None

# ════════════════════════════════════════════════════════════════
# Credential Encryption
# ════════════════════════════════════════════════════════════════
_ENCRYPT_KEY_FILE = os.environ.get("ENCRYPT_KEY_FILE", "/app/.encrypt.key")

def _get_fernet():
    """Get or create the Fernet encryption key."""
    if os.path.exists(_ENCRYPT_KEY_FILE):
        key = open(_ENCRYPT_KEY_FILE, "rb").read().strip()
    else:
        key = Fernet.generate_key()
        os.makedirs(os.path.dirname(_ENCRYPT_KEY_FILE), exist_ok=True)
        with open(_ENCRYPT_KEY_FILE, "wb") as f:
            f.write(key)
        log.info("Generated new encryption key")
    return Fernet(key)

def encrypt_value(plaintext: str) -> str:
    """Encrypt a string, return base64-encoded ciphertext prefixed with 'enc:'."""
    if not plaintext:
        return ""
    return "enc:" + _get_fernet().encrypt(plaintext.encode()).decode()

def decrypt_value(stored: str) -> str:
    """Decrypt a value. Returns as-is if not encrypted (backwards compat)."""
    if not stored:
        return ""
    if stored.startswith("enc:"):
        try:
            return _get_fernet().decrypt(stored[4:].encode()).decode()
        except Exception:
            log.warning("Failed to decrypt value — returning empty")
            return ""
    return stored  # plain text (legacy)

# ════════════════════════════════════════════════════════════════
# Source priority: agent(50) > winrm(40) > scvmm(30) > vcenter(20) > sccm(10)
# ════════════════════════════════════════════════════════════════
SOURCE_PRIORITY = {"agent": 50, "winrm": 40, "scvmm": 30, "vcenter": 20, "sccm": 10}


# ════════════════════════════════════════════════════════════════
# Startup: auto-migrate + launch collector
# ════════════════════════════════════════════════════════════════
async def _run_migrations():
    """Run all SQL migration files in order."""
    migration_dir = os.environ.get("MIGRATION_DIR", "/app/migrations")
    if not os.path.isdir(migration_dir):
        migration_dir = os.path.join(os.path.dirname(__file__), "..", "db", "migrations")
    if not os.path.isdir(migration_dir):
        log.warning(f"No migration directory found at {migration_dir}")
        return
    conn = await asyncpg.connect(host=DB_HOST, port=DB_PORT, database=DB_NAME, user=DB_USER, password=DB_PASS)
    try:
        await conn.execute("CREATE TABLE IF NOT EXISTS _migrations (filename TEXT PRIMARY KEY, applied_at TIMESTAMPTZ DEFAULT NOW())")
        applied = {r["filename"] for r in await conn.fetch("SELECT filename FROM _migrations")}
        files = sorted(glob.glob(os.path.join(migration_dir, "*.sql")))
        for f in files:
            name = os.path.basename(f)
            if name not in applied:
                log.info(f"Applying migration: {name}")
                sql = open(f).read()
                await conn.execute(sql)
                await conn.execute("INSERT INTO _migrations (filename) VALUES ($1)", name)
                log.info(f"  ✓ {name}")
    finally:
        await conn.close()


@asynccontextmanager
async def lifespan(app: FastAPI):
    global pool
    await _run_migrations()
    pool = await asyncpg.create_pool(host=DB_HOST, port=DB_PORT, database=DB_NAME, user=DB_USER, password=DB_PASS, min_size=2, max_size=10)
    log.info("Database pool ready")
    yield
    await pool.close()


app = FastAPI(title="MS License True-Up", version="2.0.0", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


@app.middleware("http")
async def _no_cache_api(request: Request, call_next):
    """Prevent browsers from serving stale /api responses (e.g. dashboard not refreshing)."""
    response = await call_next(request)
    if request.url.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


# ════════════════════════════════════════════════════════════════
# Smart Upsert — THE one function used everywhere
# ════════════════════════════════════════════════════════════════
async def upsert_host(conn, data: dict, scan_source: str = "winrm") -> int:
    """Per-field smart merge. Higher-priority source wins for fields it provides.
    Lower-priority source fills NULLs only. is_virtual=TRUE always wins.
    User overrides (license_override, sql_license_override) never touched."""
    hostname = str(data.get("hostname", "")).strip()
    if "." in hostname:
        hostname = hostname.split(".")[0].upper()
    else:
        hostname = hostname.upper()

    new_prio = SOURCE_PRIORITY.get(scan_source, 0)
    existing = await conn.fetchrow("SELECT id, scan_source FROM hosts WHERE hostname=$1", hostname)
    existing_prio = SOURCE_PRIORITY.get((existing["scan_source"] or "") if existing else "", 0)
    is_higher = new_prio >= existing_prio

    return await conn.fetchval("""
        INSERT INTO hosts (hostname, ip_address, domain, os_name, os_version, os_edition,
            is_virtual, hypervisor_host, cpu_sockets, cpu_cores, cpu_logical, ram_gb,
            cpu_model, last_scan, scan_source)
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,NOW(),$14)
        ON CONFLICT (hostname) DO UPDATE SET
            ip_address = CASE WHEN $15 THEN COALESCE(NULLIF(EXCLUDED.ip_address,''), hosts.ip_address)
                              ELSE COALESCE(NULLIF(hosts.ip_address,''), EXCLUDED.ip_address) END,
            domain = CASE WHEN $15 THEN COALESCE(NULLIF(EXCLUDED.domain,''), hosts.domain)
                          ELSE COALESCE(NULLIF(hosts.domain,''), EXCLUDED.domain) END,
            os_name = CASE WHEN $15 THEN COALESCE(NULLIF(EXCLUDED.os_name,''), hosts.os_name)
                           ELSE COALESCE(NULLIF(hosts.os_name,''), EXCLUDED.os_name) END,
            os_version = CASE WHEN $15 THEN COALESCE(NULLIF(EXCLUDED.os_version,''), hosts.os_version)
                              ELSE COALESCE(NULLIF(hosts.os_version,''), EXCLUDED.os_version) END,
            os_edition = CASE WHEN $15 THEN COALESCE(NULLIF(EXCLUDED.os_edition,''), hosts.os_edition)
                              ELSE COALESCE(NULLIF(hosts.os_edition,''), EXCLUDED.os_edition) END,
            is_virtual = CASE WHEN EXCLUDED.is_virtual = TRUE THEN TRUE
                              ELSE COALESCE(hosts.is_virtual, EXCLUDED.is_virtual) END,
            hypervisor_host = CASE WHEN $15 THEN COALESCE(NULLIF(EXCLUDED.hypervisor_host,''), hosts.hypervisor_host)
                                   ELSE COALESCE(NULLIF(hosts.hypervisor_host,''), EXCLUDED.hypervisor_host) END,
            cpu_sockets = CASE WHEN $15 THEN COALESCE(EXCLUDED.cpu_sockets, hosts.cpu_sockets)
                               ELSE COALESCE(hosts.cpu_sockets, EXCLUDED.cpu_sockets) END,
            cpu_cores = CASE WHEN $15 THEN COALESCE(EXCLUDED.cpu_cores, hosts.cpu_cores)
                             ELSE COALESCE(hosts.cpu_cores, EXCLUDED.cpu_cores) END,
            cpu_logical = CASE WHEN $15 THEN COALESCE(EXCLUDED.cpu_logical, hosts.cpu_logical)
                               ELSE COALESCE(hosts.cpu_logical, EXCLUDED.cpu_logical) END,
            ram_gb = CASE WHEN $15 THEN COALESCE(EXCLUDED.ram_gb, hosts.ram_gb)
                          ELSE COALESCE(hosts.ram_gb, EXCLUDED.ram_gb) END,
            cpu_model = CASE WHEN $15 THEN COALESCE(NULLIF(EXCLUDED.cpu_model,''), hosts.cpu_model)
                             ELSE COALESCE(NULLIF(hosts.cpu_model,''), EXCLUDED.cpu_model) END,
            last_scan = CASE WHEN $15 THEN NOW() ELSE hosts.last_scan END,
            scan_source = CASE WHEN $15 THEN $14 ELSE hosts.scan_source END,
            updated_at = NOW()
        RETURNING id
    """, hostname,
        str(data.get("ip_address")) if data.get("ip_address") else None,
        data.get("domain"), data.get("os_name"), data.get("os_version"), data.get("os_edition"),
        bool(data.get("is_virtual", False)), data.get("hypervisor_host"),
        _safe_int(data.get("cpu_sockets")), _safe_int(data.get("cpu_cores")),
        _safe_int(data.get("cpu_logical")), _safe_float(data.get("ram_gb")),
        data.get("cpu_model"), scan_source, is_higher)


async def upsert_sql_instance(conn, host_id: int, data: dict):
    await conn.execute("""
        INSERT INTO sql_instances (host_id, instance_name, edition, version, version_name,
            license_model, is_clustered, cluster_name, last_scan)
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,NOW())
        ON CONFLICT (host_id, instance_name) DO UPDATE SET
            edition=COALESCE(EXCLUDED.edition, sql_instances.edition),
            version=COALESCE(EXCLUDED.version, sql_instances.version),
            version_name=COALESCE(EXCLUDED.version_name, sql_instances.version_name),
            license_model=COALESCE(EXCLUDED.license_model, sql_instances.license_model),
            is_clustered=EXCLUDED.is_clustered,
            cluster_name=COALESCE(EXCLUDED.cluster_name, sql_instances.cluster_name),
            last_scan=NOW()
    """, host_id, data["instance_name"], data.get("edition"), data.get("version"),
        data.get("version_name"), data.get("license_model"),
        data.get("is_clustered", False), data.get("cluster_name"))


async def upsert_product(conn, host_id: int, data: dict):
    await conn.execute("""
        INSERT INTO installed_products (host_id, product_name, product_family,
            version, edition, last_scan)
        VALUES ($1,$2,$3,$4,$5,NOW())
        ON CONFLICT (host_id, product_name) DO UPDATE SET
            product_family=COALESCE(EXCLUDED.product_family, installed_products.product_family),
            version=COALESCE(EXCLUDED.version, installed_products.version),
            edition=COALESCE(EXCLUDED.edition, installed_products.edition),
            last_scan=NOW()
    """, host_id, data["product_name"], data.get("product_family"),
        data.get("version"), data.get("edition"))


def _safe_int(v):
    if v is None: return None
    try: return int(v)
    except (ValueError, TypeError): return None

def _safe_float(v):
    if v is None: return None
    try: return float(v)
    except (ValueError, TypeError): return None


# ════════════════════════════════════════════════════════════════
# Unified License Calculation — THE one function for everything
# ════════════════════════════════════════════════════════════════
def _is_windows_server(os_name):
    n = (os_name or "").lower()
    return "server" in n and "esxi" not in n

def _eff_edition(h):
    return h.get("license_override") or h.get("os_edition") or "Standard"

def compute_license_assignments(all_hosts):
    """Compute license assignments for every host. Used by hosts page, compliance, dashboard, export."""
    # Build Datacenter physical host set
    dc_hostnames = set()
    for h in all_hosts:
        if not h["is_virtual"] and h["hostname"]:
            ed = _eff_edition(h)
            if ed == "Datacenter":
                dc_hostnames.add(h["hostname"].upper())
                dc_hostnames.add(h["hostname"].upper().split(".")[0])

    # Build SQL Enterprise physical host set
    sql_ent_hostnames = set()
    for h in all_hosts:
        if not h["is_virtual"] and h["hostname"] and h.get("sql_license_override") == "Enterprise":
            sql_ent_hostnames.add(h["hostname"].upper())
            sql_ent_hostnames.add(h["hostname"].upper().split(".")[0])

    for h in all_hosts:
        override = h.get("license_override")
        is_server = _is_windows_server(h.get("os_name"))

        # ── Windows Server license ──
        if not h["is_virtual"]:
            if override == "None":
                h["license_assignment"] = "No License"
                h["licensed_cores"] = 0
            elif override == "Vendor":
                h["license_assignment"] = "Vendor Provided"
                h["licensed_cores"] = 0
            elif override or is_server:
                ed = override or h.get("os_edition") or "Standard"
                sockets = h["cpu_sockets"] or 1
                cores = h["cpu_cores"] or 0
                cps = max(cores // sockets if sockets else cores, 8) if cores else 8
                host_cores = max(cps * sockets, 16)
                h["license_assignment"] = f"{ed} ({host_cores}c)"
                h["licensed_cores"] = host_cores
            else:
                h["license_assignment"] = "N/A"
                h["licensed_cores"] = None
        else:
            if override == "None":
                h["license_assignment"] = "No License"
                h["licensed_cores"] = 0
            elif override == "Vendor":
                h["license_assignment"] = "Vendor Provided"
                h["licensed_cores"] = 0
            else:
                hyp = (h.get("hypervisor_host") or "").upper()
                hyp_short = hyp.split(".")[0] if hyp else ""
                if hyp in dc_hostnames or hyp_short in dc_hostnames:
                    h["license_assignment"] = "DC Covered"
                    h["licensed_cores"] = 0
                elif not is_server and not override:
                    h["license_assignment"] = "N/A"
                    h["licensed_cores"] = None
                else:
                    ed = override or h.get("os_edition") or "Standard"
                    h["license_assignment"] = f"{ed} (VM)"
                    h["licensed_cores"] = None

        # ── SQL license ──
        sql_override = h.get("sql_license_override")
        if sql_override == "None":
            h["sql_license_assignment"] = "—"
        elif sql_override == "Vendor":
            h["sql_license_assignment"] = "Vendor Provided"
        elif sql_override in ("Enterprise", "Standard") and not h["is_virtual"]:
            sockets = h["cpu_sockets"] or 1
            cores = h["cpu_cores"] or 0
            cps = max(cores // sockets if sockets else cores, 4) if cores else 4
            sql_cores = cps * sockets
            label = "Ent" if sql_override == "Enterprise" else "Std"
            h["sql_license_assignment"] = f"SQL {label} ({sql_cores}c)"
        elif h["is_virtual"]:
            hyp = (h.get("hypervisor_host") or "").upper()
            hyp_short = hyp.split(".")[0] if hyp else ""
            if hyp in sql_ent_hostnames or hyp_short in sql_ent_hostnames:
                h["sql_license_assignment"] = "SQL Ent Covered"
            else:
                h["sql_license_assignment"] = "—"
        else:
            h["sql_license_assignment"] = "—"


# ════════════════════════════════════════════════════════════════
# API Endpoints
# ════════════════════════════════════════════════════════════════

# ─── Dashboard ───
@app.get("/api/dashboard")
async def dashboard():
    async with pool.acquire() as conn:
        hosts = await conn.fetchval("SELECT COUNT(*) FROM hosts WHERE status='active'")
        vms = await conn.fetchval("SELECT COUNT(*) FROM hosts WHERE status='active' AND is_virtual=TRUE")
        sql_count = await conn.fetchval("SELECT COUNT(*) FROM sql_instances")
        products = await conn.fetchval("SELECT COUNT(DISTINCT product_family) FROM installed_products")
        last_scan = await conn.fetchrow("SELECT * FROM scan_log ORDER BY id DESC LIMIT 1")
        # Source breakdown
        source_rows = await conn.fetch("SELECT scan_source, COUNT(*) as cnt FROM hosts WHERE status='active' GROUP BY scan_source ORDER BY cnt DESC")
        sources = {r["scan_source"]: r["cnt"] for r in source_rows}
        # OS breakdown
        os_rows = await conn.fetch("""
            SELECT CASE
                WHEN os_name ILIKE '%esxi%' THEN 'VMware ESXi'
                WHEN os_name ILIKE '%windows server 2022%' THEN 'Windows Server 2022'
                WHEN os_name ILIKE '%windows server 2019%' THEN 'Windows Server 2019'
                WHEN os_name ILIKE '%windows server 2016%' THEN 'Windows Server 2016'
                WHEN os_name ILIKE '%windows server 2012%' THEN 'Windows Server 2012'
                WHEN os_name ILIKE '%windows server%' THEN 'Windows Server (Other)'
                WHEN os_name ILIKE '%red hat%' OR os_name ILIKE '%rhel%' THEN 'Red Hat Linux'
                WHEN os_name ILIKE '%ubuntu%' THEN 'Ubuntu Linux'
                WHEN os_name ILIKE '%suse%' THEN 'SUSE Linux'
                WHEN os_name ILIKE '%linux%' THEN 'Linux (Other)'
                WHEN os_name IS NULL OR os_name = '' THEN 'Unknown'
                ELSE 'Other'
            END as os_group, COUNT(*) as cnt
            FROM hosts WHERE status='active' GROUP BY os_group ORDER BY cnt DESC
        """)
        os_breakdown = {r["os_group"]: r["cnt"] for r in os_rows}
        return {
            "total_hosts": hosts, "physical_hosts": hosts - vms, "virtual_hosts": vms,
            "sql_instances": sql_count, "product_families": products,
            "last_scan": dict(last_scan) if last_scan else None,
            "sources": sources, "os_breakdown": os_breakdown,
        }


# ─── API Key Management ───
def _hash_key(key: str) -> str:
    return hashlib.sha256(key.encode()).hexdigest()

async def _verify_api_key(key: str):
    key_hash = _hash_key(key)
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, status, expires_at FROM api_keys WHERE key_hash=$1", key_hash)
        if not row:
            # Fallback: check legacy single-key setting
            stored = await conn.fetchval("SELECT value FROM settings WHERE key='agent_api_key'")
            if stored and decrypt_value(stored) == key:
                return
            raise HTTPException(401, "Invalid API key")
        if row["status"] != "active":
            raise HTTPException(401, "API key is expired or revoked")
        if row["expires_at"] and row["expires_at"] < datetime.now(timezone.utc):
            await conn.execute("UPDATE api_keys SET status='expired' WHERE id=$1", row["id"])
            raise HTTPException(401, "API key has expired")
        # Track usage
        await conn.execute(
            "UPDATE api_keys SET last_used_at=NOW(), use_count=use_count+1 WHERE id=$1", row["id"])

@app.get("/api/api-keys")
async def list_api_keys():
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT id, name, description, key_prefix, status, expires_at, last_used_at, use_count, created_at FROM api_keys ORDER BY created_at DESC")
        return {"api_keys": [dict(r) for r in rows]}

@app.post("/api/api-keys")
async def create_api_key(data: dict):
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(400, "Name is required")
    description = data.get("description", "")
    expires_at = data.get("expires_at")  # ISO string or null

    key = secrets.token_urlsafe(32)
    key_hash = _hash_key(key)
    key_prefix = key[:8]

    exp = None
    if expires_at:
        try:
            exp = datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
        except ValueError:
            raise HTTPException(400, "Invalid expires_at format")

    async with pool.acquire() as conn:
        kid = await conn.fetchval(
            "INSERT INTO api_keys (name, description, key_hash, key_prefix, expires_at) VALUES ($1,$2,$3,$4,$5) RETURNING id",
            name, description, key_hash, key_prefix, exp)
    return {"id": kid, "key": key, "prefix": key_prefix}

@app.put("/api/api-keys/{kid}")
async def update_api_key(kid: int, data: dict):
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE api_keys SET name=$2, description=$3 WHERE id=$1",
            kid, data.get("name", ""), data.get("description", ""))
    return {"status": "updated"}

@app.post("/api/api-keys/{kid}/expire")
async def expire_api_key(kid: int):
    async with pool.acquire() as conn:
        await conn.execute("UPDATE api_keys SET status='expired', expires_at=NOW() WHERE id=$1", kid)
    return {"status": "expired"}

@app.delete("/api/api-keys/{kid}")
async def delete_api_key(kid: int):
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM api_keys WHERE id=$1", kid)
    return {"status": "deleted"}

# ─── Agent Push Ingest ───
@app.post("/api/ingest")
async def ingest(data: dict, x_api_key: str = Header(...)):
    await _verify_api_key(x_api_key)
    host_data = data.get("host")
    if not host_data or not host_data.get("hostname"):
        raise HTTPException(400, "host.hostname is required")
    scan_source = data.get("scan_source", "agent")
    if scan_source not in SOURCE_PRIORITY:
        scan_source = "agent"

    async with pool.acquire() as conn:
        host_id = await upsert_host(conn, host_data, scan_source)
        for inst in (data.get("sql_instances") or []):
            if inst.get("instance_name"):
                await upsert_sql_instance(conn, host_id, inst)
        for prod in (data.get("installed_products") or []):
            if prod.get("product_name"):
                await upsert_product(conn, host_id, prod)
    return {"status": "ok", "hostname": host_data["hostname"], "host_id": host_id}


@app.post("/api/settings/generate-api-key")
async def generate_api_key():
    """Legacy endpoint — kept for backward compatibility."""
    key = secrets.token_urlsafe(32)
    async with pool.acquire() as conn:
        await conn.execute("UPDATE settings SET value=$1, updated_at=NOW() WHERE key='agent_api_key'", encrypt_value(key))
    return {"key": key}


# ─── Hosts ───
@app.get("/api/hosts")
async def list_hosts(status: str = "active", page: int = 1, per_page: int = 50):
    offset = (page - 1) * per_page
    async with pool.acquire() as conn:
        total = await conn.fetchval("SELECT COUNT(*) FROM hosts WHERE status=$1", status)
        rows = await conn.fetch("SELECT * FROM hosts WHERE status=$1 ORDER BY hostname", status)
        all_hosts = [dict(r) for r in rows]
        compute_license_assignments(all_hosts)
        return {"total": total, "page": page, "per_page": per_page, "hosts": all_hosts[offset:offset+per_page]}


@app.patch("/api/hosts/{host_id}/license")
async def set_license(host_id: int, data: dict):
    override = data.get("license_override")
    if override is not None and override not in ("Datacenter", "Standard", "None", "Vendor"):
        raise HTTPException(400, "Invalid license_override")
    async with pool.acquire() as conn:
        await conn.execute("UPDATE hosts SET license_override=$1, updated_at=NOW() WHERE id=$2", override, host_id)
    return {"status": "ok", "host_id": host_id, "license_override": override}


@app.patch("/api/hosts/bulk-license")
async def bulk_license(data: dict):
    host_ids = data.get("host_ids", [])
    override = data.get("license_override")
    if not host_ids:
        raise HTTPException(400, "host_ids required")
    async with pool.acquire() as conn:
        await conn.execute("UPDATE hosts SET license_override=$1, updated_at=NOW() WHERE id = ANY($2::int[])", override, host_ids)
    return {"status": "ok", "updated": len(host_ids)}


@app.patch("/api/hosts/{host_id}/sql-license")
async def set_sql_license(host_id: int, data: dict):
    override = data.get("sql_license_override")
    if override is not None and override not in ("Enterprise", "Standard", "None", "Vendor"):
        raise HTTPException(400, "Invalid sql_license_override")
    async with pool.acquire() as conn:
        await conn.execute("UPDATE hosts SET sql_license_override=$1, updated_at=NOW() WHERE id=$2", override, host_id)
    return {"status": "ok", "host_id": host_id}


@app.patch("/api/hosts/bulk-sql-license")
async def bulk_sql_license(data: dict):
    host_ids = data.get("host_ids", [])
    override = data.get("sql_license_override")
    if not host_ids:
        raise HTTPException(400, "host_ids required")
    async with pool.acquire() as conn:
        await conn.execute("UPDATE hosts SET sql_license_override=$1, updated_at=NOW() WHERE id = ANY($2::int[])", override, host_ids)
    return {"status": "ok", "updated": len(host_ids)}


@app.get("/api/hosts/inactive")
async def list_inactive():
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT id, hostname, ip_address, os_name, scan_source, last_scan FROM hosts WHERE status='inactive' ORDER BY last_scan DESC NULLS LAST")
        return {"hosts": [dict(r) for r in rows]}


@app.patch("/api/hosts/reactivate")
async def reactivate(data: dict):
    host_ids = data.get("host_ids", [])
    if not host_ids: raise HTTPException(400, "host_ids required")
    async with pool.acquire() as conn:
        await conn.execute("UPDATE hosts SET status='active', updated_at=NOW() WHERE id = ANY($1::int[]) AND status='inactive'", host_ids)
    return {"status": "ok", "reactivated": len(host_ids)}


@app.delete("/api/hosts/bulk-delete")
async def bulk_delete(data: dict):
    host_ids = data.get("host_ids", [])
    if not host_ids: raise HTTPException(400, "host_ids required")
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM sql_instances WHERE host_id = ANY($1::int[])", host_ids)
        await conn.execute("DELETE FROM installed_products WHERE host_id = ANY($1::int[])", host_ids)
        await conn.execute("DELETE FROM hosts WHERE id = ANY($1::int[])", host_ids)
    return {"status": "ok", "deleted": len(host_ids)}


@app.delete("/api/hosts/{host_id}")
async def delete_host(host_id: int):
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM sql_instances WHERE host_id=$1", host_id)
        await conn.execute("DELETE FROM installed_products WHERE host_id=$1", host_id)
        await conn.execute("DELETE FROM hosts WHERE id=$1", host_id)
    return {"status": "ok"}


# ─── Compliance ───
@app.get("/api/compliance")
async def compliance_report():
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM hosts WHERE status='active' ORDER BY hostname")
        all_hosts = [dict(r) for r in rows]
        compute_license_assignments(all_hosts)
        gaps = []

        # Build physical/VM sets using the same logic as compute_license_assignments
        dc_hostnames = set()
        std_hostnames = set()
        for h in all_hosts:
            if not h["is_virtual"] and h["hostname"]:
                ed = _eff_edition(h)
                name_upper = h["hostname"].upper()
                if ed == "Datacenter":
                    dc_hostnames.add(name_upper)
                    dc_hostnames.add(name_upper.split(".")[0])
                elif ed == "Standard":
                    std_hostnames.add(name_upper)
                    std_hostnames.add(name_upper.split(".")[0])

        # Count Windows VMs per hypervisor host
        win_vms_per_hyp = {}
        total_vms_per_hyp = {}
        for h in all_hosts:
            if h["is_virtual"] and h.get("hypervisor_host"):
                hyp = h["hypervisor_host"].upper()
                hyp_short = hyp.split(".")[0]
                total_vms_per_hyp[hyp] = total_vms_per_hyp.get(hyp, 0) + 1
                total_vms_per_hyp[hyp_short] = total_vms_per_hyp.get(hyp_short, 0) + 1
                if _is_windows_server(h.get("os_name")):
                    win_vms_per_hyp[hyp] = win_vms_per_hyp.get(hyp, 0) + 1
                    win_vms_per_hyp[hyp_short] = win_vms_per_hyp.get(hyp_short, 0) + 1

        # Classify hosts
        ws_hosts = [h for h in all_hosts
                    if _is_windows_server(h["os_name"])
                    or (h.get("license_override") and h["license_override"] not in ("None", "Vendor"))]

        physical_by_ed = {}
        all_vms = []
        for h in ws_hosts:
            if h.get("license_override") in ("None", "Vendor"):
                continue
            ed = _eff_edition(h)
            if not h["is_virtual"]:
                physical_by_ed.setdefault(ed, []).append(h)
            else:
                all_vms.append(h)

        # Classify VMs — track ALL VMs per hypervisor for stacking math
        covered_dc_count = 0
        covered_dc_vms = []
        std_vms_by_hyp = {}       # hyp_key → [vm, vm, ...]
        uncovered_by_ed = {}
        excluded_vms = []
        orphan_vms = []           # VMs with no known hypervisor host
        for vm in all_vms:
            if vm.get("license_override") in ("None", "Vendor"):
                excluded_vms.append(vm)
                continue
            hyp = (vm.get("hypervisor_host") or "").upper()
            hyp_short = hyp.split(".")[0] if hyp else ""
            if hyp in dc_hostnames or hyp_short in dc_hostnames:
                covered_dc_count += 1
                covered_dc_vms.append(vm)
            elif hyp_short in std_hostnames or hyp in std_hostnames:
                key = hyp_short or hyp
                std_vms_by_hyp.setdefault(key, []).append(vm)
            elif hyp:
                # VM on a host we don't recognize as physical — needs its own license
                ed = _eff_edition(vm)
                uncovered_by_ed.setdefault(ed, []).append(vm)
            else:
                orphan_vms.append(vm)
                ed = _eff_edition(vm)
                uncovered_by_ed.setdefault(ed, []).append(vm)

        all_editions = set(physical_by_ed.keys()) | set(uncovered_by_ed.keys())
        # Also include Standard if we have std VMs
        if std_vms_by_hyp:
            all_editions.add("Standard")

        for edition in sorted(all_editions):
            total_cores = 0
            host_details = []
            hyp_summary = []  # per-hypervisor stacking breakdown

            # Physical hosts
            for h in physical_by_ed.get(edition, []):
                sockets = h["cpu_sockets"] or 1
                cores = h["cpu_cores"] or 0
                cps = max(cores // sockets if sockets else cores, 8) if cores else 8
                hc = max(cps * sockets, 16)
                total_cores += hc
                hname_upper = h["hostname"].upper()
                wvms = win_vms_per_hyp.get(hname_upper, 0) or win_vms_per_hyp.get(hname_upper.split(".")[0], 0)
                tvms = total_vms_per_hyp.get(hname_upper, 0) or total_vms_per_hyp.get(hname_upper.split(".")[0], 0)
                host_details.append({"hostname": h["hostname"], "sockets": sockets,
                    "physical_cores": cores, "licensed_cores": hc,
                    "two_core_packs": math.ceil(hc / 2), "type": "physical",
                    "status": "needs_license", "reason": "Physical host — 1 base license required",
                    "windows_vm_count": wvms, "total_vm_count": tvms})

            # Standard stacking: group VMs by hypervisor
            if edition == "Standard":
                for hyp_key, vms in sorted(std_vms_by_hyp.items()):
                    vm_count = len(vms)
                    licenses_needed = math.ceil(vm_count / 2)  # each license covers 2 VMs
                    extra_licenses = max(0, licenses_needed - 1)  # minus the 1 base license already counted above
                    extra_cores = extra_licenses * 16
                    total_cores += extra_cores

                    hyp_summary.append({
                        "hypervisor": hyp_key, "vm_count": vm_count,
                        "windows_vm_count": win_vms_per_hyp.get(hyp_key, 0),
                        "total_vm_count": total_vms_per_hyp.get(hyp_key, 0),
                        "licenses_needed": licenses_needed,
                        "base_included": 1, "extra_stacked": extra_licenses,
                        "extra_cores": extra_cores
                    })

                    for i, vm in enumerate(vms):
                        slot = i + 1
                        license_num = math.ceil(slot / 2)
                        is_covered = license_num <= 1  # first 2 VMs covered by base license
                        host_details.append({
                            "hostname": vm["hostname"], "sockets": vm["cpu_sockets"] or 1,
                            "physical_cores": vm["cpu_cores"] or 0,
                            "licensed_cores": 0 if is_covered else 0,
                            "two_core_packs": 0, "type": "vm",
                            "status": "covered" if is_covered else "needs_license",
                            "reason": f"VM {slot}/{vm_count} on {hyp_key} — license {license_num} of {licenses_needed}" + (" (base)" if license_num == 1 else " (stacked)"),
                            "hypervisor_host": hyp_key,
                            "license_num": license_num,
                            "stacking_group": hyp_key
                        })

            # Datacenter covered VMs
            if edition == "Datacenter":
                for vm in covered_dc_vms:
                    host_details.append({"hostname": vm["hostname"], "sockets": vm["cpu_sockets"] or 1,
                        "physical_cores": vm["cpu_cores"] or 0, "licensed_cores": 0,
                        "two_core_packs": 0, "type": "vm",
                        "status": "covered", "reason": f"Covered by DC host ({vm.get('hypervisor_host','')})",
                        "hypervisor_host": vm.get("hypervisor_host", "")})

            # Orphan/unknown VMs needing license
            uncovered = uncovered_by_ed.get(edition, [])
            if edition != "Datacenter" and uncovered:
                extra = math.ceil(len(uncovered) / 2) * 16
                total_cores += extra
            for vm in uncovered:
                hyp = vm.get("hypervisor_host") or "Unknown"
                host_details.append({"hostname": vm["hostname"], "sockets": vm["cpu_sockets"] or 1,
                    "physical_cores": vm["cpu_cores"] or 0, "licensed_cores": 16,
                    "two_core_packs": 8, "type": "vm",
                    "status": "needs_license", "reason": f"VM on unknown host ({hyp}) — needs own license",
                    "hypervisor_host": hyp})

            packs = math.ceil(total_cores / 2)
            entitled = await conn.fetchval("""
                SELECT COALESCE(SUM(CASE WHEN e.license_type='core_2pack' THEN e.quantity*2
                    WHEN e.license_type='core' THEN e.quantity ELSE 0 END), 0)
                FROM entitlements e
                LEFT JOIN agreements a ON a.id = e.agreement_id
                WHERE e.product_family='WindowsServer' AND e.edition=$1
                  AND (e.agreement_id IS NULL OR a.expiry_date IS NULL OR a.expiry_date >= CURRENT_DATE)
            """, edition)

            note = ""
            total_std_vms = sum(len(v) for v in std_vms_by_hyp.values()) if edition == "Standard" else 0
            if edition == "Datacenter":
                note = f"Covers unlimited VMs ({covered_dc_count} covered)"
            elif edition == "Standard" and std_vms_by_hyp:
                total_stacked = sum(max(0, math.ceil(len(v)/2) - 1) for v in std_vms_by_hyp.values())
                note = f"{total_std_vms} VMs across {len(std_vms_by_hyp)} hosts · {len(physical_by_ed.get(edition,[]))} base + {total_stacked} stacked licenses"
            elif uncovered:
                note = f"{len(uncovered)} VMs need additional licenses"

            gaps.append({
                "product": f"Windows Server {edition}", "license_type": "core",
                "physical_hosts": len(physical_by_ed.get(edition, [])),
                "virtual_hosts": total_std_vms if edition == "Standard" else (covered_dc_count if edition == "Datacenter" else len(uncovered)),
                "required_cores": total_cores, "required_2packs": packs,
                "entitled_cores": entitled, "entitled_2packs": entitled // 2 if entitled else 0,
                "gap_cores": max(0, total_cores - entitled),
                "gap_2packs": max(0, packs - (entitled // 2 if entitled else 0)),
                "compliant": entitled >= total_cores, "host_details": host_details, "note": note,
                "hyp_summary": hyp_summary if edition == "Standard" else [],
            })

        # ── SQL Server ──
        sql_hosts = await conn.fetch("""
            SELECT si.edition, si.instance_name, h.hostname, h.cpu_sockets, h.cpu_cores,
                   h.is_virtual, h.hypervisor_host
            FROM sql_instances si JOIN hosts h ON h.id=si.host_id
            WHERE h.status='active' AND si.license_model='core'
              AND COALESCE(h.sql_license_override, '') NOT IN ('None', 'Vendor')
        """)
        sql_by_ed = {}
        for r in sql_hosts:
            sql_by_ed.setdefault(r["edition"] or "Standard", []).append(r)

        override_hosts = await conn.fetch("""
            SELECT sql_license_override as edition, hostname, cpu_sockets, cpu_cores,
                   is_virtual, hypervisor_host
            FROM hosts WHERE status='active' AND sql_license_override IS NOT NULL
              AND sql_license_override NOT IN ('None', 'Vendor')
        """)
        override_names = set()
        for r in override_hosts:
            ed = r["edition"]
            override_names.add(r["hostname"])
            sql_by_ed.setdefault(ed, [])
            if not any(x["hostname"] == r["hostname"] for x in sql_by_ed[ed]):
                sql_by_ed[ed].append(dict(r))

        sql_ent_phys = set()
        for ed, insts in sql_by_ed.items():
            if "enterprise" in ed.lower():
                for i in insts:
                    if not i.get("is_virtual"):
                        sql_ent_phys.add((i["hostname"] or "").upper())

        for edition, instances in sql_by_ed.items():
            total_cores = 0
            seen = set()
            details = []
            phys_count = vm_count = 0
            for inst in instances:
                hk = inst["hostname"]
                if hk in seen: continue
                seen.add(hk)
                is_vm = bool(inst.get("is_virtual"))
                sockets = inst.get("cpu_sockets") or 1
                cores = inst.get("cpu_cores") or 0
                hyp = (inst.get("hypervisor_host") or "").upper()
                hyp_short = hyp.split(".")[0] if hyp else ""
                covered = is_vm and (hyp in sql_ent_phys or hyp_short in sql_ent_phys)
                if is_vm: vm_count += 1
                else: phys_count += 1
                if not covered:
                    cps = max(cores // sockets if sockets else cores, 4) if cores else 4
                    hc = cps * sockets
                    total_cores += hc
                else:
                    hc = 0
                details.append({"hostname": hk, "sockets": sockets, "physical_cores": cores,
                    "licensed_cores": hc, "two_core_packs": math.ceil(hc/2),
                    "type": "vm" if is_vm else "physical",
                    "source": "override" if hk in override_names else "discovered",
                    "covered_by_enterprise": covered,
                    "status": "covered" if covered else "needs_license",
                    "reason": "Covered by Enterprise physical host" if covered else ("Physical host" if not is_vm else "VM needs own SQL license"),
                    "hypervisor_host": inst.get("hypervisor_host", "")})

            packs = math.ceil(total_cores / 2)
            entitled = await conn.fetchval("""
                SELECT COALESCE(SUM(CASE WHEN e.license_type='core_2pack' THEN e.quantity*2
                    WHEN e.license_type='core' THEN e.quantity ELSE 0 END), 0)
                FROM entitlements e
                LEFT JOIN agreements a ON a.id = e.agreement_id
                WHERE e.product_family='SQLServer' AND e.edition ILIKE $1
                  AND (e.agreement_id IS NULL OR a.expiry_date IS NULL OR a.expiry_date >= CURRENT_DATE)
            """, f"%{edition.split()[0]}%")

            gaps.append({
                "product": f"SQL Server {edition}", "license_type": "core",
                "physical_hosts": phys_count, "virtual_hosts": vm_count,
                "hosts": len(seen), "instances": len(instances),
                "required_cores": total_cores, "required_2packs": packs,
                "entitled_cores": entitled, "entitled_2packs": entitled // 2 if entitled else 0,
                "gap_cores": max(0, total_cores - entitled),
                "gap_2packs": max(0, packs - (entitled // 2 if entitled else 0)),
                "compliant": entitled >= total_cores, "host_details": details,
            })

        return {"compliance": gaps}


# ─── License Summaries ───
@app.get("/api/licenses/windows-server")
async def ws_licenses():
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT os_edition, is_virtual, COUNT(*) as host_count,
                   SUM(cpu_cores) as total_cores, SUM(cpu_sockets) as total_sockets
            FROM hosts WHERE status='active' AND os_name ILIKE '%server%'
            GROUP BY os_edition, is_virtual ORDER BY os_edition
        """)
        ent = await conn.fetch("""
            SELECT e.edition, e.license_type, SUM(e.quantity) as quantity FROM entitlements e
            LEFT JOIN agreements a ON a.id = e.agreement_id
            WHERE e.product_family='WindowsServer'
              AND (e.agreement_id IS NULL OR a.expiry_date IS NULL OR a.expiry_date >= CURRENT_DATE)
            GROUP BY e.edition, e.license_type
        """)
        return {"discovered": [dict(r) for r in rows], "entitlements": [dict(r) for r in ent]}


@app.get("/api/licenses/sql-server")
async def sql_licenses():
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT si.edition, si.version_name, si.license_model, COUNT(*) as instance_count, SUM(h.cpu_cores) as total_host_cores
            FROM sql_instances si JOIN hosts h ON h.id=si.host_id WHERE h.status='active'
            GROUP BY si.edition, si.version_name, si.license_model ORDER BY si.edition
        """)
        ent = await conn.fetch("""
            SELECT e.edition, e.license_type, SUM(e.quantity) as quantity FROM entitlements e
            LEFT JOIN agreements a ON a.id = e.agreement_id
            WHERE e.product_family='SQLServer'
              AND (e.agreement_id IS NULL OR a.expiry_date IS NULL OR a.expiry_date >= CURRENT_DATE)
            GROUP BY e.edition, e.license_type
        """)
        return {"discovered": [dict(r) for r in rows], "entitlements": [dict(r) for r in ent]}


@app.get("/api/licenses/products")
async def product_summary():
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT product_family, product_name, edition, COUNT(*) as install_count
            FROM installed_products ip JOIN hosts h ON h.id=ip.host_id WHERE h.status='active'
            GROUP BY product_family, product_name, edition ORDER BY product_family
        """)
        return {"products": [dict(r) for r in rows]}


# ─── Entitlements CRUD ───
# ─── Agreements ───
@app.get("/api/agreements")
async def list_agreements():
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM agreements ORDER BY expiry_date DESC NULLS LAST")
        return {"agreements": [dict(r) for r in rows]}

@app.post("/api/agreements")
async def create_agreement(data: dict):
    async with pool.acquire() as conn:
        aid = await conn.fetchval("""
            INSERT INTO agreements (name, agreement_number, agreement_type, start_date, expiry_date, notes)
            VALUES ($1,$2,$3,$4,$5,$6) RETURNING id
        """, data["name"], data.get("agreement_number") or None, data.get("agreement_type", "EA"),
            _parse_date(data.get("start_date")), _parse_date(data.get("expiry_date")), data.get("notes") or None)
    return {"status": "created", "id": aid}

@app.put("/api/agreements/{aid}")
async def update_agreement(aid: int, data: dict):
    async with pool.acquire() as conn:
        await conn.execute("""
            UPDATE agreements SET name=$2, agreement_number=$3, agreement_type=$4,
                start_date=$5, expiry_date=$6, notes=$7 WHERE id=$1
        """, aid, data["name"], data.get("agreement_number") or None, data.get("agreement_type", "EA"),
            _parse_date(data.get("start_date")), _parse_date(data.get("expiry_date")), data.get("notes") or None)
    return {"status": "updated"}

@app.delete("/api/agreements/{aid}")
async def delete_agreement(aid: int):
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM agreements WHERE id=$1", aid)
    return {"status": "deleted"}

# ─── Entitlements ───
@app.get("/api/entitlements")
async def list_entitlements():
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT e.*, a.name as agreement_name, a.expiry_date as agreement_expiry,
                   a.agreement_number as agr_number
            FROM entitlements e
            LEFT JOIN agreements a ON a.id = e.agreement_id
            ORDER BY e.product_family, e.edition
        """)
        return {"entitlements": [dict(r) for r in rows]}

@app.post("/api/entitlements")
async def create_entitlement(data: dict):
    async with pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO entitlements (product_name, product_family, edition, license_type,
                quantity, agreement_number, agreement_type, effective_date, expiry_date,
                sa_included, notes, agreement_id, part_number)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13)
        """, data["product_name"], data.get("product_family"), data.get("edition"),
            data.get("license_type"), data["quantity"],
            data.get("agreement_number"), data.get("agreement_type"),
            data.get("effective_date"), data.get("expiry_date"),
            data.get("sa_included", False), data.get("notes"),
            data.get("agreement_id"), data.get("part_number"))
    return {"status": "created"}

@app.put("/api/entitlements/{ent_id}")
async def update_entitlement(ent_id: int, data: dict):
    async with pool.acquire() as conn:
        await conn.execute("""
            UPDATE entitlements SET product_name=$2, product_family=$3, edition=$4,
                license_type=$5, quantity=$6, agreement_number=$7, agreement_type=$8,
                effective_date=$9, expiry_date=$10, sa_included=$11, notes=$12,
                agreement_id=$13, part_number=$14 WHERE id=$1
        """, ent_id, data["product_name"], data.get("product_family"), data.get("edition"),
            data.get("license_type"), data["quantity"],
            data.get("agreement_number"), data.get("agreement_type"),
            data.get("effective_date"), data.get("expiry_date"),
            data.get("sa_included", False), data.get("notes"),
            data.get("agreement_id"), data.get("part_number"))
    return {"status": "updated"}

@app.delete("/api/entitlements/{ent_id}")
async def delete_entitlement(ent_id: int):
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM entitlements WHERE id=$1", ent_id)
    return {"status": "deleted"}


# ─── Scan History ───
@app.get("/api/scans")
async def scan_history(limit: int = 20):
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM scan_log ORDER BY id DESC LIMIT $1", limit)
        return {"scans": [dict(r) for r in rows]}

@app.post("/api/scans/trigger")
async def trigger_scan(data: dict = None):
    scanners = (data or {}).get("scanners", ["all"])
    async with pool.acquire() as conn:
        scan_id = await conn.fetchval(
            "INSERT INTO scan_log (scan_type, status) VALUES ($1, 'running') RETURNING id",
            f"manual:{','.join(scanners)}")
    # Run the scan in the background immediately
    asyncio.create_task(_run_manual_scan(scan_id))
    return {"status": "triggered", "scan_id": scan_id, "scanners": scanners}


# ─── Settings CRUD ───
@app.get("/api/settings")
async def get_settings(category: Optional[str] = None):
    async with pool.acquire() as conn:
        if category:
            rows = await conn.fetch("SELECT key, value, category, description, sensitive FROM settings WHERE category=$1 ORDER BY key", category)
        else:
            rows = await conn.fetch("SELECT key, value, category, description, sensitive FROM settings ORDER BY category, key")
        result = []
        for r in rows:
            d = dict(r)
            if d["sensitive"] and d["value"]:
                d["value"] = "••••••••"
            result.append(d)
        return {"settings": result}

SENSITIVE_SETTINGS = {"winrm_password", "vcenter_password", "sccm_password", "agent_api_key", "snmp_community"}

@app.put("/api/settings")
async def update_settings(data: dict):
    settings = data.get("settings", {})
    if not settings: raise HTTPException(400, "No settings provided")
    async with pool.acquire() as conn:
        for key, value in settings.items():
            if value == "••••••••": continue
            store_value = encrypt_value(str(value)) if key in SENSITIVE_SETTINGS else str(value)
            await conn.execute("UPDATE settings SET value=$2, updated_at=NOW() WHERE key=$1", key, store_value)
    return {"status": "updated"}

@app.get("/api/settings/raw/{key}")
async def get_setting_raw(key: str):
    """Return decrypted setting value (for internal use by scanners)."""
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT value, sensitive FROM settings WHERE key=$1", key)
        if not row: raise HTTPException(404, f"Setting '{key}' not found")
        value = decrypt_value(row["value"]) if row["sensitive"] else row["value"]
        return {"key": key, "value": value}


# ─── Targets CRUD ───
@app.get("/api/targets")
async def list_targets(scan_type: Optional[str] = None):
    async with pool.acquire() as conn:
        if scan_type:
            rows = await conn.fetch("SELECT * FROM targets WHERE scan_type=$1 ORDER BY hostname", scan_type)
        else:
            rows = await conn.fetch("SELECT * FROM targets ORDER BY scan_type, hostname")
        return {"targets": [dict(r) for r in rows]}

@app.post("/api/targets")
async def add_target(data: dict):
    hostname = data.get("hostname", "").strip()
    scan_type = data.get("scan_type", "winrm")
    if not hostname: raise HTTPException(400, "hostname required")
    is_subnet = "/" in hostname
    async with pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO targets (hostname, scan_type, enabled, notes, is_subnet, credential_id)
            VALUES ($1,$2,$3,$4,$5,$6) ON CONFLICT (hostname, scan_type) DO UPDATE SET
            enabled=EXCLUDED.enabled, notes=EXCLUDED.notes, credential_id=EXCLUDED.credential_id
        """, hostname, scan_type, data.get("enabled", True), data.get("notes", ""), is_subnet, data.get("credential_id"))
    return {"status": "added"}

@app.post("/api/targets/bulk")
async def bulk_targets(data: dict):
    hostnames = data.get("hostnames", [])
    scan_type = data.get("scan_type", "winrm")
    async with pool.acquire() as conn:
        for h in hostnames:
            h = h.strip()
            if h and not h.startswith("#"):
                await conn.execute("INSERT INTO targets (hostname, scan_type, is_subnet) VALUES ($1,$2,$3) ON CONFLICT DO NOTHING", h, scan_type, "/" in h)
    return {"status": "added", "count": len(hostnames)}

@app.delete("/api/targets/{tid}")
async def delete_target(tid: int):
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM targets WHERE id=$1", tid)
    return {"status": "deleted"}


# ─── Credentials CRUD ───
@app.get("/api/credentials")
async def list_creds():
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM credentials ORDER BY cred_type, name")
        result = []
        for r in rows:
            d = dict(r)
            d["password"] = "********" if d.get("password") else ""
            result.append(d)
        return {"credentials": result}

@app.post("/api/credentials")
async def create_cred(data: dict):
    enc_password = encrypt_value(data.get("password", ""))
    enc_community = encrypt_value(data.get("community", ""))
    async with pool.acquire() as conn:
        cid = await conn.fetchval("""
            INSERT INTO credentials (name, cred_type, username, password, domain, transport, port, use_https, verify_ssl, community, snmp_version, notes)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12) RETURNING id
        """, data["name"], data["cred_type"], data.get("username"), enc_password,
            data.get("domain"), data.get("transport", "ntlm"), data.get("port"),
            data.get("use_https", False), data.get("verify_ssl", False),
            enc_community, data.get("snmp_version"), data.get("notes"))
    return {"status": "created", "id": cid}

@app.put("/api/credentials/{cid}")
async def update_cred(cid: int, data: dict):
    async with pool.acquire() as conn:
        if data.get("password") == "********":
            await conn.execute("""
                UPDATE credentials SET name=$2, username=$3, domain=$4, transport=$5, port=$6,
                    use_https=$7, verify_ssl=$8, community=$9, snmp_version=$10, notes=$11, enabled=$12 WHERE id=$1
            """, cid, data["name"], data.get("username"), data.get("domain"),
                data.get("transport"), data.get("port"), data.get("use_https", False),
                data.get("verify_ssl", False),
                encrypt_value(data.get("community", "")) if data.get("community") else None,
                data.get("snmp_version"),
                data.get("notes"), data.get("enabled", True))
        else:
            await conn.execute("""
                UPDATE credentials SET name=$2, username=$3, password=$4, domain=$5, transport=$6, port=$7,
                    use_https=$8, verify_ssl=$9, community=$10, snmp_version=$11, notes=$12, enabled=$13 WHERE id=$1
            """, cid, data["name"], data.get("username"), encrypt_value(data.get("password", "")),
                data.get("domain"),
                data.get("transport"), data.get("port"), data.get("use_https", False),
                data.get("verify_ssl", False),
                encrypt_value(data.get("community", "")) if data.get("community") else None,
                data.get("snmp_version"),
                data.get("notes"), data.get("enabled", True))
    return {"status": "updated"}

@app.delete("/api/credentials/{cid}")
async def delete_cred(cid: int):
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM credentials WHERE id=$1", cid)
    return {"status": "deleted"}

@app.get("/api/credentials/{cid}/raw")
async def get_cred_raw(cid: int):
    """Return credential with password decrypted (for edit form & scanners)."""
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM credentials WHERE id=$1", cid)
        if not row: raise HTTPException(404, "Not found")
        d = dict(row)
        d["password"] = decrypt_value(d.get("password", ""))
        d["community"] = decrypt_value(d.get("community", ""))
        return d


# ─── vCenter Instances ───
@app.get("/api/vcenter-instances")
async def list_vcenters():
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT v.*, c.name as credential_name, c.username
            FROM vcenter_instances v LEFT JOIN credentials c ON c.id=v.credential_id ORDER BY v.name
        """)
        return {"instances": [dict(r) for r in rows]}

@app.post("/api/vcenter-instances")
async def create_vcenter(data: dict):
    async with pool.acquire() as conn:
        vid = await conn.fetchval("INSERT INTO vcenter_instances (name, hostname, credential_id, enabled, notes) VALUES ($1,$2,$3,$4,$5) RETURNING id",
            data["name"], data["hostname"], data.get("credential_id"), data.get("enabled", True), data.get("notes"))
    return {"status": "created", "id": vid}

@app.put("/api/vcenter-instances/{vid}")
async def update_vcenter(vid: int, data: dict):
    async with pool.acquire() as conn:
        await conn.execute("UPDATE vcenter_instances SET name=$2, hostname=$3, credential_id=$4, enabled=$5, notes=$6 WHERE id=$1",
            vid, data["name"], data["hostname"], data.get("credential_id"), data.get("enabled", True), data.get("notes"))
    return {"status": "updated"}

@app.delete("/api/vcenter-instances/{vid}")
async def delete_vcenter(vid: int):
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM vcenter_instances WHERE id=$1", vid)
    return {"status": "deleted"}


# ─── SCCM Instances ───
@app.get("/api/sccm-instances")
async def list_sccm():
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT s.*, c.name as credential_name, c.username
            FROM sccm_instances s LEFT JOIN credentials c ON c.id=s.credential_id ORDER BY s.name
        """)
        return {"instances": [dict(r) for r in rows]}

@app.post("/api/sccm-instances")
async def create_sccm(data: dict):
    async with pool.acquire() as conn:
        sid = await conn.fetchval(
            "INSERT INTO sccm_instances (name, server_url, credential_id, verify_ssl, enabled, notes) VALUES ($1,$2,$3,$4,$5,$6) RETURNING id",
            data["name"], data["server_url"], data.get("credential_id"), data.get("verify_ssl", False), data.get("enabled", True), data.get("notes"))
    return {"status": "created", "id": sid}

@app.put("/api/sccm-instances/{sid}")
async def update_sccm(sid: int, data: dict):
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE sccm_instances SET name=$2, server_url=$3, credential_id=$4, verify_ssl=$5, enabled=$6, notes=$7 WHERE id=$1",
            sid, data["name"], data["server_url"], data.get("credential_id"), data.get("verify_ssl", False), data.get("enabled", True), data.get("notes"))
    return {"status": "updated"}

@app.delete("/api/sccm-instances/{sid}")
async def delete_sccm_instance(sid: int):
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM sccm_instances WHERE id=$1", sid)
    return {"status": "deleted"}


# ─── ESXi → Hyper-V Migration Tracking ───
_MIGRATION_FIELDS = {
    "vm_name", "power_state", "datastore", "app_support", "schedule_jeremy",
    "move_daytime", "move_afterhours", "scheduled_at", "migrated", "migrated_override",
    "date_migrated", "verified_working", "zprl_to_nrep", "vpg_deleted",
    "should_be_zprl", "excluded", "notes",
}

def _coerce_migration_fields(fields: dict) -> dict:
    """Convert date/datetime strings from the UI into proper objects for asyncpg."""
    if "date_migrated" in fields:
        dv = fields["date_migrated"]
        if not dv:
            fields["date_migrated"] = None
        elif isinstance(dv, str):
            try:
                fields["date_migrated"] = date.fromisoformat(dv[:10])
            except ValueError:
                raise HTTPException(400, "date_migrated must be YYYY-MM-DD")
    if "scheduled_at" in fields:
        sv = fields["scheduled_at"]
        if not sv:
            fields["scheduled_at"] = None
        elif isinstance(sv, str):
            try:
                fields["scheduled_at"] = datetime.fromisoformat(sv)
            except ValueError:
                raise HTTPException(400, "scheduled_at must be ISO datetime")
    return fields

@app.get("/api/migrations")
async def list_migrations():
    async with pool.acquire() as conn:
        # VM names currently discovered in SCVMM (i.e., now running on Hyper-V)
        scvmm_rows = await conn.fetch(
            "SELECT DISTINCT hostname FROM hosts WHERE status='active' AND scan_source='scvmm'")
        scvmm_last = await conn.fetchval(
            "SELECT MAX(last_scan) FROM hosts WHERE status='active' AND scan_source='scvmm'")
        scvmm = set()
        for r in scvmm_rows:
            h = (r["hostname"] or "").upper()
            if h:
                scvmm.add(h)
                scvmm.add(h.split(".")[0])

        def detected(name):
            if not name:
                return False
            u = name.upper()
            return u in scvmm or u.split(".")[0] in scvmm

        rows = await conn.fetch("SELECT * FROM migration_tracking ORDER BY migrated ASC, vm_name ASC")
        items = [dict(r) for r in rows]

        # Auto-flag VMs detected in SCVMM as migrated, unless manually overridden or excluded
        to_flip = [r["id"] for r in items
                   if not r["migrated_override"] and not r["migrated"]
                   and not r["excluded"] and detected(r["vm_name"])]
        if to_flip:
            await conn.execute("""
                UPDATE migration_tracking
                SET migrated=TRUE, date_migrated=COALESCE(date_migrated, CURRENT_DATE), updated_at=NOW()
                WHERE id = ANY($1::int[])
            """, to_flip)
            rows = await conn.fetch("SELECT * FROM migration_tracking ORDER BY migrated ASC, vm_name ASC")
            items = [dict(r) for r in rows]

    for r in items:
        r["detected_scvmm"] = detected(r["vm_name"])
    # Progress counts ignore excluded VMs (those that won't be migrated)
    active = [r for r in items if not r["excluded"]]
    total = len(active)
    done = sum(1 for r in active if r["migrated"])
    excluded = sum(1 for r in items if r["excluded"])
    return {"migrations": items, "total": total, "migrated": done,
            "remaining": total - done, "excluded": excluded,
            "scvmm_count": len(scvmm_rows), "scvmm_last_seen": scvmm_last}

@app.post("/api/migrations")
async def create_migration(data: dict):
    vm_name = (data.get("vm_name") or "").strip()
    if not vm_name:
        raise HTTPException(400, "vm_name required")
    async with pool.acquire() as conn:
        try:
            mid = await conn.fetchval(
                "INSERT INTO migration_tracking (vm_name, power_state, datastore, notes) VALUES ($1,$2,$3,$4) RETURNING id",
                vm_name, data.get("power_state"), data.get("datastore"), data.get("notes"))
        except asyncpg.UniqueViolationError:
            raise HTTPException(409, f"{vm_name} is already tracked")
    return {"status": "created", "id": mid}

@app.patch("/api/migrations/{mid}")
async def update_migration(mid: int, data: dict):
    # Only allow known columns; build a dynamic SET clause
    fields = {k: v for k, v in data.items() if k in _MIGRATION_FIELDS}
    if not fields:
        raise HTTPException(400, "No valid fields to update")
    _coerce_migration_fields(fields)
    cols = list(fields.keys())
    set_clause = ", ".join(f"{c}=${i+2}" for i, c in enumerate(cols))
    values = [fields[c] for c in cols]
    async with pool.acquire() as conn:
        await conn.execute(
            f"UPDATE migration_tracking SET {set_clause}, updated_at=NOW() WHERE id=$1",
            mid, *values)
    return {"status": "updated"}

@app.delete("/api/migrations/{mid}")
async def delete_migration(mid: int):
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM migration_tracking WHERE id=$1", mid)
    return {"status": "deleted"}

@app.post("/api/migrations/bulk")
async def bulk_update_migrations(data: dict):
    """Apply the same field(s) to many rows at once (e.g. exclude/include)."""
    ids = data.get("ids", [])
    fields = {k: v for k, v in (data.get("fields") or {}).items() if k in _MIGRATION_FIELDS}
    if not ids or not fields:
        raise HTTPException(400, "ids and fields are required")
    _coerce_migration_fields(fields)
    cols = list(fields.keys())
    set_clause = ", ".join(f"{c}=${i+2}" for i, c in enumerate(cols))
    values = [fields[c] for c in cols]
    async with pool.acquire() as conn:
        await conn.execute(
            f"UPDATE migration_tracking SET {set_clause}, updated_at=NOW() WHERE id = ANY($1::int[])",
            ids, *values)
    return {"status": "ok", "updated": len(ids)}

@app.post("/api/migrations/bulk-delete")
async def bulk_delete_migrations(data: dict):
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(400, "ids required")
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM migration_tracking WHERE id = ANY($1::int[])", ids)
    return {"status": "ok", "deleted": len(ids)}

@app.post("/api/migrations/import-esxi")
async def import_esxi_vms():
    """Add any active vCenter/ESXi VMs that aren't already in the tracker."""
    async with pool.acquire() as conn:
        added = await conn.fetch("""
            INSERT INTO migration_tracking (vm_name)
            SELECT hostname FROM hosts
            WHERE status='active' AND is_virtual=TRUE AND scan_source='vcenter'
            ON CONFLICT (vm_name) DO NOTHING
            RETURNING id
        """)
    return {"status": "ok", "added": len(added)}


# ─── Excel Export ───
@app.get("/api/export/trueup")
async def export_excel():
    from openpyxl import Workbook
    wb = Workbook()

    def _clean(rec):
        # openpyxl can't write timezone-aware datetimes; strip tzinfo (values are stored UTC)
        out = []
        for v in rec.values():
            if isinstance(v, datetime) and v.tzinfo is not None:
                v = v.replace(tzinfo=None)
            out.append(v)
        return out

    async with pool.acquire() as conn:
        ws_rows = await conn.fetch("SELECT hostname, os_edition, CASE WHEN is_virtual THEN 'Virtual' ELSE 'Physical' END as type, cpu_sockets, cpu_cores, domain, last_scan FROM hosts WHERE status='active' AND os_name ILIKE '%server%' ORDER BY os_edition, hostname")
        sql_rows = await conn.fetch("SELECT h.hostname, si.instance_name, si.edition, si.version_name, si.license_model, h.cpu_cores, si.is_clustered FROM sql_instances si JOIN hosts h ON h.id=si.host_id WHERE h.status='active' ORDER BY si.edition")
        prod_rows = await conn.fetch("SELECT h.hostname, ip.product_name, ip.product_family, ip.version, ip.edition FROM installed_products ip JOIN hosts h ON h.id=ip.host_id WHERE h.status='active' ORDER BY ip.product_family")

    compliance = await compliance_report()

    ws1 = wb.active
    ws1.title = "Windows Server"
    ws1.append(["Hostname", "Edition", "Type", "Sockets", "Cores", "Domain", "Last Scan"])
    for r in ws_rows: ws1.append(_clean(r))

    ws2 = wb.create_sheet("SQL Server")
    ws2.append(["Hostname", "Instance", "Edition", "Version", "License Model", "Cores", "Clustered"])
    for r in sql_rows: ws2.append(_clean(r))

    ws3 = wb.create_sheet("Other Products")
    ws3.append(["Hostname", "Product", "Family", "Version", "Edition"])
    for r in prod_rows: ws3.append(_clean(r))

    ws4 = wb.create_sheet("Compliance")
    ws4.append(["Product", "Type", "Required Cores", "Required 2-Packs", "Entitled Cores", "Entitled 2-Packs", "Gap", "Compliant", "Notes"])
    for item in compliance.get("compliance", []):
        ws4.append([item.get("product"), item.get("license_type"), item.get("required_cores", 0),
            item.get("required_2packs", 0), item.get("entitled_cores", 0), item.get("entitled_2packs", 0),
            item.get("gap_2packs", 0), "Yes" if item.get("compliant") else "NO", item.get("note", "")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=trueup_report.xlsx"})


# ─── Collector Restart ───
@app.post("/api/collector/restart")
async def restart_collector():
    async with pool.acquire() as conn:
        await conn.execute(
            "INSERT INTO settings (key, value, category, description) VALUES ('_collector_restart', $1, 'system', 'Restart signal') ON CONFLICT (key) DO UPDATE SET value=$1",
            str(__import__('time').time()))
    return {"status": "restart_signaled"}


# ════════════════════════════════════════════════════════════════
# Git-based Update System
# ════════════════════════════════════════════════════════════════
GIT_DIR = os.environ.get("GIT_REPO_PATH", "/app/.git")

def _git(cmd, cwd=None):
    """Run a git command and return output."""
    work_tree = os.path.dirname(GIT_DIR) if GIT_DIR.endswith(".git") else GIT_DIR
    git_dir = GIT_DIR if GIT_DIR.endswith(".git") else os.path.join(GIT_DIR, ".git")
    full_cmd = f"git --git-dir={git_dir} --work-tree={work_tree} {cmd}"
    try:
        result = subprocess.run(full_cmd, shell=True, capture_output=True, text=True, timeout=60)
        return {"ok": result.returncode == 0, "stdout": result.stdout.strip(), "stderr": result.stderr.strip()}
    except Exception as e:
        return {"ok": False, "stdout": "", "stderr": str(e)}


@app.get("/api/system/version")
async def get_version():
    """Get current git version info."""
    head = _git("rev-parse --short HEAD")
    branch = _git("rev-parse --abbrev-ref HEAD")
    log_result = _git("log -1 --format=%s")
    remote_url = _git("remote get-url origin")
    return {
        "current_hash": head["stdout"] if head["ok"] else "unknown",
        "branch": branch["stdout"] if branch["ok"] else "unknown",
        "last_commit": log_result["stdout"] if log_result["ok"] else "unknown",
        "remote": remote_url["stdout"] if remote_url["ok"] else "not configured",
    }


@app.get("/api/system/check-updates")
async def check_updates():
    """Fetch from remote and check if updates are available."""
    fetch = _git("fetch origin")
    if not fetch["ok"]:
        return {"available": False, "error": fetch["stderr"]}
    head = _git("rev-parse HEAD")
    remote = _git("rev-parse @{u}")
    if not remote["ok"]:
        return {"available": False, "error": "No upstream tracking branch"}
    if head["stdout"] == remote["stdout"]:
        return {"available": False, "message": "Already up to date"}
    # Get list of changes
    changes = _git("log --oneline HEAD..@{u}")
    return {
        "available": True,
        "current": head["stdout"][:8],
        "latest": remote["stdout"][:8],
        "changes": changes["stdout"].split("\n") if changes["ok"] else [],
    }


@app.post("/api/system/update")
async def apply_update():
    """Pull latest changes, rebuild and restart the app container."""
    # Pull from git (the .git volume is the host repo)
    _git("stash")
    pull = _git("pull --ff-only")
    if not pull["ok"]:
        _git("stash pop")
        return {"status": "error", "message": pull["stderr"]}
    head = _git("rev-parse --short HEAD")
    # Log the update
    async with pool.acquire() as conn:
        await conn.execute(
            "INSERT INTO app_versions (git_hash, git_branch, notes) VALUES ($1, $2, $3)",
            head["stdout"], _git("rev-parse --abbrev-ref HEAD")["stdout"], pull["stdout"])
    # Trigger rebuild via Docker socket in background
    # The host repo was already updated by git pull above (shared .git volume),
    # so we shell out to docker compose on the host via the socket
    asyncio.create_task(_rebuild_container())
    return {
        "status": "updated",
        "new_version": head["stdout"],
        "output": pull["stdout"],
        "rebuilding": True,
        "message": "Pulling and rebuilding — the app will restart in ~30 seconds.",
    }


async def _rebuild_container():
    """Write a trigger file that the host-side update-watcher.sh picks up
    and runs 'docker compose up -d --build app'."""
    await asyncio.sleep(2)  # Let the HTTP response go out first
    log.info("Writing rebuild trigger...")
    try:
        # .git is mounted from the host repo root, so ../. is the project dir
        trigger = Path("/app/.git") / ".." / ".rebuild-trigger"
        trigger.resolve().write_text(str(datetime.now(timezone.utc)))
        log.info("Rebuild trigger written — watcher will rebuild shortly")
    except Exception as e:
        log.error(f"Failed to write trigger: {e}")


@app.get("/api/system/update-history")
async def update_history():
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM app_versions ORDER BY updated_at DESC LIMIT 20")
        return {"history": [dict(r) for r in rows]}


# ─── Script Downloads ───
SCRIPTS_DIR = Path(__file__).parent / "scripts"
if not SCRIPTS_DIR.is_dir():
    SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"

SCRIPT_CATALOG = [
    {
        "id": "agent",
        "name": "TrueUp Agent",
        "filename": "TrueUp-Agent.ps1",
        "description": "Runs locally on each Windows Server as a scheduled task. Collects OS, CPU, SQL, and product inventory then POSTs to the API. No WinRM or firewall rules needed.",
        "install": "Deploy via GPO or run the Deploy script. Creates a daily scheduled task running as SYSTEM.",
    },
    {
        "id": "deploy-agent",
        "name": "Agent Deploy Script",
        "filename": "Deploy-TrueUpAgent.ps1",
        "description": "Installs TrueUp-Agent.ps1 as a daily scheduled task on the local server. Idempotent — safe to run multiple times.",
        "install": "Run as Administrator on each server, or push via GPO startup script.",
    },
    {
        "id": "scvmm",
        "name": "SCVMM Collector",
        "filename": "TrueUp-SCVMM.ps1",
        "description": "Queries System Center VMM for all Hyper-V hosts and VMs, then pushes inventory to the API. Run on the SCVMM server.",
        "install": "Schedule as a daily task on your SCVMM management server. Requires the VirtualMachineManager PowerShell module.",
    },
    {
        "id": "enable-winrm",
        "name": "Enable WinRM",
        "filename": "enable-winrm.ps1",
        "description": "Enables and configures WinRM on target servers for remote scanning.",
        "install": "Run as Administrator on target servers or push via GPO.",
    },
]

@app.get("/api/scripts")
async def list_scripts():
    """Return catalog of available scripts with availability status."""
    result = []
    for s in SCRIPT_CATALOG:
        info = dict(s)
        info["available"] = (SCRIPTS_DIR / s["filename"]).is_file()
        result.append(info)
    return {"scripts": result}

@app.get("/api/scripts/{script_id}/download")
async def download_script(script_id: str, request: Request, key: str = Query(default="")):
    """Download a script with server URL and API key pre-configured."""
    entry = next((s for s in SCRIPT_CATALOG if s["id"] == script_id), None)
    if not entry:
        raise HTTPException(404, "Script not found")
    script_path = SCRIPTS_DIR / entry["filename"]
    if not script_path.is_file():
        raise HTTPException(404, f"Script file not found: {entry['filename']}")

    content = script_path.read_text(encoding="utf-8")

    # Derive server URL from the incoming request
    host_header = request.headers.get("host", "")
    scheme = request.headers.get("x-forwarded-proto", "http")
    server_url = f"{scheme}://{host_header}" if host_header else "http://localhost:3000"

    # Replace ApiUrl wherever it appears as a param default
    content = re.sub(
        r'(\[string\]\$ApiUrl\s*=\s*)"[^"]*"',
        rf'\g<1>"{server_url}"',
        content
    )
    # Replace ApiKey
    if key:
        content = re.sub(
            r'(\[string\]\$ApiKey\s*=\s*)"[^"]*"',
            rf'\g<1>"{key}"',
            content
        )
    else:
        content = re.sub(
            r'(\[string\]\$ApiKey\s*=\s*)"[^"]*"',
            r'\g<1>"PASTE_YOUR_API_KEY_HERE"',
            content
        )

    return PlainTextResponse(
        content=content,
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{entry["filename"]}"'}
    )


# ─── Logs ───
_log_buffer: list[str] = []
_LOG_BUFFER_MAX = 2000

class _BufferedLogHandler(logging.Handler):
    """Captures log records into an in-memory ring buffer for the UI."""
    def emit(self, record):
        line = self.format(record)
        _log_buffer.append(line)
        if len(_log_buffer) > _LOG_BUFFER_MAX:
            del _log_buffer[:len(_log_buffer) - _LOG_BUFFER_MAX]

# Install the handler on startup
_buf_handler = _BufferedLogHandler()
_buf_handler.setLevel(logging.DEBUG)
_buf_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
logging.getLogger("trueup").addHandler(_buf_handler)

@app.get("/api/logs")
async def get_logs(lines: int = 200, level: str = ""):
    """Return recent application log lines."""
    result = _log_buffer[-lines:]
    if level:
        level = level.upper()
        result = [l for l in result if f"[{level}]" in l]
    return {"logs": result, "total": len(_log_buffer)}

@app.get("/api/scans/{scan_id}/errors")
async def scan_errors(scan_id: int):
    """Return detailed errors for a specific scan."""
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT * FROM scan_errors WHERE scan_id=$1 ORDER BY created_at", scan_id)
        return {"errors": [dict(r) for r in rows]}


# ════════════════════════════════════════════════════════════════
# Serve React frontend (must be last — catches all non-API routes)
# ════════════════════════════════════════════════════════════════
static_dir = Path(__file__).parent / "static"
if static_dir.is_dir():
    app.mount("/assets", StaticFiles(directory=static_dir / "assets"), name="assets") if (static_dir / "assets").is_dir() else None

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        """Serve index.html for all non-API routes (React SPA routing)."""
        file_path = static_dir / full_path
        if full_path and file_path.is_file():
            return FileResponse(file_path)
        return FileResponse(static_dir / "index.html")


# ════════════════════════════════════════════════════════════════
# Background Collector — runs scanners on a schedule
# ════════════════════════════════════════════════════════════════
async def _get_setting(conn, key, default=""):
    row = await conn.fetchrow("SELECT value, sensitive FROM settings WHERE key=$1", key)
    if not row: return default
    return decrypt_value(row["value"]) if row["sensitive"] else row["value"]


async def _mark_stale_hosts(conn, stale_days=30):
    """Mark hosts as inactive if not scanned recently.
    Agent and SCVMM are PUSH-based — never mark them stale."""
    await conn.execute("""
        UPDATE hosts SET status='inactive', updated_at=NOW()
        WHERE status='active' AND last_scan < NOW() - INTERVAL '1 day' * $1
          AND scan_source NOT IN ('agent', 'scvmm')
    """, stale_days)


async def _run_manual_scan(scan_id: int):
    """Execute a scan immediately for a manually triggered scan_id."""
    try:
        await _execute_scanners(scan_id)
    except Exception as e:
        log.error(f"Manual scan error: {e}")
        async with pool.acquire() as conn:
            await conn.execute("UPDATE scan_log SET status='error', error_message=$1 WHERE id=$2", str(e)[:1000], scan_id)


async def _run_scan_cycle():
    """Run all enabled scanners on the scheduled interval."""
    log.info("Starting scan cycle")
    async with pool.acquire() as conn:
        scan_id = await conn.fetchval(
            "INSERT INTO scan_log (scan_type, status) VALUES ('scheduled', 'running') RETURNING id")
    await _execute_scanners(scan_id)


async def _execute_scanners(scan_id: int):
    """Core scanner execution shared by manual and scheduled scans."""
    log.info(f"Scan {scan_id}: starting scanners")
    total_scanned = 0
    total_failed = 0
    all_errors = []

    try:
        async with pool.acquire() as conn:
            # ── WinRM ──
            winrm_enabled = await _get_setting(conn, "winrm_enabled", "false")
            if winrm_enabled.lower() == "true":
                targets = await conn.fetch("SELECT * FROM targets WHERE scan_type='winrm' AND enabled=TRUE")
                if targets:
                    cred = {
                        "username": await _get_setting(conn, "winrm_username"),
                        "password": await _get_setting(conn, "winrm_password"),
                        "transport": await _get_setting(conn, "winrm_transport", "ntlm"),
                        "port": int(await _get_setting(conn, "winrm_port", "5985")),
                    }
                    # Check for per-target credentials
                    cred_ids = set(t["credential_id"] for t in targets if t.get("credential_id"))
                    cred_map = {}
                    if cred_ids:
                        cred_rows = await conn.fetch("SELECT * FROM credentials WHERE id = ANY($1::int[]) AND enabled=TRUE", list(cred_ids))
                        for r in cred_rows:
                            cr = dict(r)
                            cr["password"] = decrypt_value(cr.get("password", ""))
                            cr["community"] = decrypt_value(cr.get("community", ""))
                            cred_map[cr["id"]] = cr

                    from scanners import winrm as winrm_scanner
                    s, f, e = await winrm_scanner.scan(pool, [dict(t) for t in targets], cred)
                    total_scanned += s; total_failed += f; all_errors.extend(e)

            # ── SCCM ──
            instances = await conn.fetch("SELECT * FROM sccm_instances WHERE enabled=TRUE")
            if instances:
                inst_list = []
                for inst in instances:
                    d = dict(inst)
                    if d.get("credential_id"):
                        cred_row = await conn.fetchrow("SELECT * FROM credentials WHERE id=$1", d["credential_id"])
                        if cred_row:
                            cr = dict(cred_row)
                            cr["password"] = decrypt_value(cr.get("password", ""))
                            d["_credential"] = cr
                        else:
                            d["_credential"] = {}
                    else:
                        d["_credential"] = {}
                    log.info(f"SCCM {d.get('hostname', d.get('server_url', '?'))}: user={d['_credential'].get('username')}, domain={d['_credential'].get('domain')}, has_password={bool(d['_credential'].get('password'))}")
                    inst_list.append(d)
                from scanners import sccm as sccm_scanner
                s, f, e = await sccm_scanner.scan(pool, inst_list)
                total_scanned += s; total_failed += f; all_errors.extend(e)

            # ── vCenter ──
            instances = await conn.fetch("SELECT * FROM vcenter_instances WHERE enabled=TRUE")
            if instances:
                inst_list = []
                for inst in instances:
                    d = dict(inst)
                    if d.get("credential_id"):
                        cred_row = await conn.fetchrow("SELECT * FROM credentials WHERE id=$1", d["credential_id"])
                        if cred_row:
                            cr = dict(cred_row)
                            cr["password"] = decrypt_value(cr.get("password", ""))
                            d["_credential"] = cr
                        else:
                            d["_credential"] = {}
                    else:
                        d["_credential"] = {
                            "username": await _get_setting(conn, "vcenter_user"),
                            "password": await _get_setting(conn, "vcenter_password"),
                            "port": int(await _get_setting(conn, "vcenter_port", "443")),
                            "verify_ssl": await _get_setting(conn, "vcenter_verify_ssl", "false") == "true",
                        }
                    log.info(f"vCenter {d['hostname']}: user={d['_credential'].get('username')}, has_password={bool(d['_credential'].get('password'))}, port={d['_credential'].get('port')}")
                    inst_list.append(d)
                from scanners import vcenter as vcenter_scanner
                s, f, e = await vcenter_scanner.scan(pool, inst_list)
                total_scanned += s; total_failed += f; all_errors.extend(e)

            # ── SNMP ──
            snmp_enabled = await _get_setting(conn, "snmp_enabled", "false")
            if snmp_enabled.lower() == "true":
                targets = await conn.fetch("SELECT * FROM targets WHERE scan_type='snmp' AND enabled=TRUE")
                if targets:
                    community = await _get_setting(conn, "snmp_community", "public")
                    from scanners import snmp as snmp_scanner
                    s, f, e = await snmp_scanner.scan(pool, [dict(t) for t in targets], community)
                    total_scanned += s; total_failed += f; all_errors.extend(e)

            # ── Mark stale ──
            await _mark_stale_hosts(conn)

        # Update scan log
        async with pool.acquire() as conn:
            await conn.execute("""
                UPDATE scan_log SET completed_at=NOW(), hosts_scanned=$1, hosts_failed=$2,
                    status=$3 WHERE id=$4
            """, total_scanned, total_failed,
                "completed" if total_failed == 0 else "completed_with_errors", scan_id)
            for err in all_errors[:100]:
                await conn.execute(
                    "INSERT INTO scan_errors (scan_id, hostname, error_type, error_message) VALUES ($1,$2,$3,$4)",
                    scan_id, err["hostname"], err["error_type"], err["error_message"])

        log.info(f"Scan cycle complete: {total_scanned} scanned, {total_failed} failed")
    except Exception as e:
        log.error(f"Scan cycle error: {e}")
        async with pool.acquire() as conn:
            await conn.execute("UPDATE scan_log SET status='error', error_message=$1 WHERE id=$2", str(e)[:1000], scan_id)


async def collector_loop():
    """Background loop that runs scan cycles."""
    await asyncio.sleep(10)  # Wait for startup
    # Clean up any stale pending scans from previous runs
    try:
        async with pool.acquire() as conn:
            await conn.execute("UPDATE scan_log SET status='cancelled', completed_at=NOW() WHERE status IN ('pending','running')")
    except Exception:
        pass
    while True:
        try:
            async with pool.acquire() as conn:
                interval = int(await _get_setting(conn, "scan_interval_minutes", "60"))
            await _run_scan_cycle()
            await asyncio.sleep(interval * 60)
        except asyncio.CancelledError:
            break
        except Exception as e:
            log.error(f"Collector loop error: {e}")
            await asyncio.sleep(300)


# Start collector on startup
@app.on_event("startup")
async def start_collector():
    asyncio.create_task(collector_loop())
